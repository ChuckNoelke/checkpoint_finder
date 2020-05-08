import openpyxl

import numpy as np

import matplotlib.pyplot as plt

import ruptures as rpt
# generate signal
n_samples, dim, sigma = 1000, 3, 4
n_bkps = 3  # number of breakpoints
signal, bkps = rpt.pw_constant(n_samples, dim, n_bkps, noise_std=sigma)

# open data points

print(signal[0] [1])



workbook = openpyxl.load_workbook('TemperingTestdaten_var4.xlsx')
print(type(workbook))

worksheet = workbook.get_sheet_by_name('Tabelle9')
data = np.zeros((3300,1))


for i in range(0,3300):
    data[i]= float(worksheet.cell(row=i+3, column=3).value)


#for i in range(0,3300):
#    data [i][1] = float(worksheet.cell(row=i+3, column=3).value)




print (data)
# detection
# change point detection
# change point detection
model = "l1"  # "l2", "rbf"
algo = rpt.Dynp(model=model, min_size=3, jump=50).fit(data)
my_bkps = algo.predict(n_bkps=4)

# show results
rpt.show.display(data, bkps, my_bkps, figsize=(10, 6))
plt.show()